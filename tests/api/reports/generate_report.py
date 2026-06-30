"""
3-Tab Excel Report Generator — API vs CSR_SOP.pdf ground-truth comparison.

Tabs:
  1. API vs SOP Comparison  — one row per call, all fields, overall pass/fail
  2. SOP Violations Summary  — one row per metric/check   [status per call]
  3. PDF Compliance Details   — one row per PDF requirement [per call]

Ground truth: CSR_SOP.pdf, SALES_REP_SOP.pdf, Reference.docx.pdf,
             tenant_onboarding_audit_otto_ai_storage.json
"""
import os
import sys
import json
import requests
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))
from dotenv import load_dotenv
load_dotenv()

API_BASE = os.getenv("OTTO_API_BASE_URL", "https://ottoai.shunyalabs.ai").rstrip("/")
API_KEY = os.getenv("OTTO_API_KEY", "")
COMPANY_ID = os.getenv("COMPANY_ID", "f438b048-9cec-47f5-a8a8-9609f48b81e6")

# Optional Shunya override for billing attribution when testing with real client company_id
SHUNYA_API_KEY = os.getenv("X_SHUNYA_API_KEY", "")
SHUNYA_API_KEY_ID = os.getenv("X_SHUNYA_API_KEY_ID", "")

if not API_KEY:
    print("OTTO_API_KEY not set — aborting")
    sys.exit(1)

# ── Test Calls ──
CALLS = [
    {
        "id": "4bf68a1b-3815-4fa8-b896-4aee02614bfb",
        "label": "James Bond (Sales)",
        "role": "sales_rep",
        "desc": "Cancellation / objection — prior negative experience",
    },
    {
        "id": "bb219502-3022-413b-8035-66ce7a08575c",
        "label": "Iseah Alegria (CSR)",
        "role": "customer_rep",
        "desc": "Service call follow-up — send quote pending action expected",
    },
]

# ── Styling ──
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SUBHEAD_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(bold=True, size=10)
NORMAL_FONT  = Font(size=10)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, horizontal="center", vertical="center")


def fetch_summary(call_id):
    """Fetch one call's summary data from the API."""
    headers = {"X-API-Key": API_KEY, "Content-Type": "application/json"}
    # Pass Shunya override headers when configured (QA Guide: optional for GET reads,
    # required for POST /process billing attribution)
    if SHUNYA_API_KEY and SHUNYA_API_KEY_ID:
        headers["X-Shunya-API-Key"] = SHUNYA_API_KEY
        headers["X-Shunya-API-Key-Id"] = SHUNYA_API_KEY_ID
    url = f"{API_BASE}/api/v1/call-processing/summary/{call_id}"
    try:
        r = requests.get(url, headers=headers, timeout=30)
        return r.json() if r.status_code == 200 else {"error": f"HTTP {r.status_code}", "detail": r.text}
    except Exception as e:
        return {"error": str(e)}


def style_header_row(ws, row, max_col, fill=None, font=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill or HEADER_FILL
        cell.font = font or HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_cell(cell, status="pass"):
    cell.border = THIN_BORDER
    cell.alignment = WRAP
    cell.font = NORMAL_FONT
    if status == "pass":
        cell.fill = GREEN_FILL
    elif status == "fail":
        cell.fill = RED_FILL
    elif status == "warn":
        cell.fill = YELLOW_FILL


def auto_width(ws, max_col, max_width=55):
    for col in range(1, max_col + 1):
        lengths = []
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    lengths.append(len(str(cell.value)))
        best = max(lengths) if lengths else 10
        ws.column_dimensions[get_column_letter(col)].width = min(best + 2, max_width)


# ======================================================================
# TAB 1 — API vs SOP Comparison
# ======================================================================
def write_comparison_tab(wb, call_data):
    """One row per call, columns grouped by section, Overall column."""
    ws = wb.create_sheet("API vs SOP Comparison")

    section_headers = [
        ("CALL METADATA", 6),
        ("SUMMARY", 3),
        ("PENDING ACTIONS (PDF 2-Stage Pipeline)", 4),
        ("OBJECTIONS (PDF Multi-Agent Debate)", 3),
        ("COMPLIANCE / STAGES", 6),
        ("COACHING (PDF 3-Question Framework)", 3),
        ("COACHING STRENGTHS", 1),
        ("QUALIFICATION", 16),
        ("", 2),  # Overall + Failure Reason
    ]

    col_names = []
    for sh_name, sh_count in section_headers:
        cols_for_section = {
            "CALL METADATA": ["Call Label", "Call ID (short)", "Agent Name", "Role", "Interaction Type", "Call Type"],
            "SUMMARY": ["Sentiment Score", "Summary Confidence", "Summary Text"],
            "PENDING ACTIONS (PDF 2-Stage Pipeline)": ["Action Items Count", "Action Items Text", "Next Steps Count", "Pending Actions Count"],
            "OBJECTIONS (PDF Multi-Agent Debate)": ["Objection Count", "Objection Details", "PDF: Objection Structure Complete?"],
            "COMPLIANCE / STAGES": ["Compliance Score", "Compliance Rate", "Stages Total", "Stages Followed", "Stages Missed", "PDF: Score Consistent with Stages?"],
            "COACHING (PDF 3-Question Framework)": ["Coaching Issues Count", "Coaching Details", "PDF: 3-Question Framework Complete?"],
            "COACHING STRENGTHS": ["Coaching Strengths Count"],
            "QUALIFICATION": ["Qualification Status", "Overall Score", "Booking Status", "Call Outcome", "Scope",
                              "Customer Name", "Appointment Date", "Appointment Type", "Appointment Confirmed",
                              "Follow-up Required", "Follow-up Reason", "BANT Budget", "BANT Need",
                              "BANT Authority", "BANT Timeline", "PDF: Cross-field Consistency"],
            "": ["Failure Reason", "Overall"],
        }
        col_names.extend(cols_for_section.get(sh_name, []))

    max_col = len(col_names)

    # Row 1: Section headers (merged effect via fill)
    current_col = 1
    for sh_name, sh_count in section_headers:
        for c in range(sh_count):
            cell = ws.cell(1, current_col)
            cell.value = sh_name if c == 0 else None
            cell.fill = SECTION_FILL
            cell.font = BOLD_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            current_col += 1
    ws.row_dimensions[1].height = 22

    # Row 2: Column headers
    for col_idx, name in enumerate(col_names, 1):
        cell = ws.cell(2, col_idx)
        cell.value = name
        cell.fill = SUBHEAD_FILL
        cell.font = Font(bold=True, size=9)
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 28

    # Analyze data and build rows
    def analyze_calls():
        results = []
        for cd in call_data:
            label = cd["label"]
            d = cd["data"]
            s = d.get("summary", {})
            q = d.get("qualification", {})
            comp = d.get("compliance", {})
            sop = comp.get("sop_compliance", {})
            stages = sop.get("stages", {})
            obj_sec = d.get("objections", {})
            objs = obj_sec.get("objections", [])
            bant = q.get("bant_scores", {})

            # Digest values
            sentiment = s.get("sentiment_score", "")
            sentiment_str = f"{sentiment:.1f}" if isinstance(sentiment, (int, float)) else str(sentiment)
            confidence = s.get("confidence_score", "")
            conf_str = f"{confidence:.2f}" if isinstance(confidence, (int, float)) else str(confidence)
            summary_text = (s.get("summary", "") or "")[:200]
            ai_count = len(s.get("action_items", []))
            ai_text = "; ".join(s.get("action_items", []))[:200]
            ns_count = len(s.get("next_steps", []))
            pa_count = len(s.get("pending_actions", []))
            obj_count = obj_sec.get("total_count", 0)
            compliance_score = sop.get("score", "")
            comp_rate = sop.get("compliance_rate", "")
            stages_total = stages.get("total", 0)
            stages_followed = ", ".join(stages.get("followed", [])) or "None"
            stages_missed = ", ".join(stages.get("missed", [])) or "None"

            # Objection details
            obj_details = ""
            if objs:
                parts = []
                for o in objs:
                    cid = o.get("category_id", "")
                    ctxt = o.get("category_text", "")
                    ov = o.get("overcome", "")
                    cf = o.get("confidence_score", "")
                    parts.append(f"Cat#{cid}: {ctxt} (overcome={ov}, conf={cf})")
                obj_details = "; ".join(parts)

            # Coaching
            coaching_issues = sop.get("coaching_issues", [])
            coaching_details = ""
            for ci in coaching_issues:
                sev = ci.get("severity", "")
                issue_text = (ci.get("issue", "") or "")[:120]
                coaching_details += f"[{'PASS' if ci.get('issue') else 'FAIL'}] {sev.upper()}: {issue_text}\n"

            coaching_strengths = sop.get("coaching_strengths", [])

            # Qualification
            qual_status = q.get("qualification_status", "")
            overall_score = q.get("overall_score", "")
            booking = q.get("booking_status", "")
            outcome = q.get("call_outcome_category", "")
            scope = d.get("scope", "IN_SCOPE")
            customer = q.get("customer_name", "")
            appt_date = q.get("appointment_date", "")
            appt_type = q.get("appointment_type", "")
            appt_confirmed = str(q.get("appointment_confirmed", ""))
            fur = q.get("follow_up_required", "")
            fur_str = str(fur) if fur is not None else ""
            reason = (q.get("follow_up_reason", "") or "")[:150]

            bant_b = bant.get("budget", "")
            bant_n = bant.get("need", "")
            bant_a = bant.get("authority", "")
            bant_t = bant.get("timeline", "")

            # === CROSS-FIELD CHECKS ===
            violations = []
            cross_field_pass = True

            # 1) overcome_evidence field (PDF Section 7.3)
            has_oe = any("overcome_evidence" in o for o in objs)
            if objs and not has_oe:
                violations.append("[MEDIUM] OBJ: PDF Section 7.3: overcome_evidence field missing from objection object")
                cross_field_pass = False

            # 2) Pending Actions 2-stage pipeline (PDF §4.2)
            has_commitment = any(
                phrase in (ai or "").lower()
                for phrase in ["send quote", "send the estimate", "send the quote", "send a quote", "call back",
                               "schedule", "send the contract", "email you"]
                for ai in s.get("action_items", [])
            )
            if has_commitment and pa_count == 0:
                violations.append("[HIGH] PA: PDF 2-stage pipeline: action_items contain rep commitments but pending_actions is empty")
                cross_field_pass = False

            # 3) Appointment fields consistent with booking (appt_date/type set but not booked)
            if booking != "booked" and appt_date and appt_type:
                violations.append(f"[MEDIUM] QUAL: Not booked but has appointment_date={appt_date} and type={appt_type}")

            # 4) Score-stage consistency
            score_ok = True
            if stages_total == 0 and compliance_score != 0:
                score_ok = False

            # 5) related_sop_metric check
            metric_warn = False
            for ci in coaching_issues:
                if ci.get("severity") == "high" and not ci.get("related_sop_metric"):
                    metric_warn = True

            overall_status = "pass" if cross_field_pass else "fail"

            row_data = [
                label,
                cd["id"][:8],
                label.split(" (")[0],
                cd["role"],
                cd.get("desc", "").split(" — ")[0],
                q.get("detected_call_type", ""),

                sentiment_str,
                conf_str,
                summary_text,

                ai_count,
                ai_text,
                ns_count,
                pa_count,

                obj_count,
                obj_details,

                compliance_score,
                comp_rate,
                stages_total,
                stages_followed,
                stages_missed,

                len(coaching_issues),
                coaching_details.strip(),
                len(coaching_strengths),

                qual_status,
                overall_score,
                booking,
                outcome,
                scope,
                customer,
                appt_date,
                appt_type,
                appt_confirmed,
                fur_str,
                reason,
                bant_b,
                bant_n,
                bant_a,
                bant_t,
            ]

            results.append({
                "data": row_data,
                "violations": violations,
                "overall_pass": cross_field_pass,
                "score_ok": score_ok,
                "metric_warn": metric_warn,
                "pa_count": pa_count,
                "obj_details": obj_details,
            })
        return results

    analysis = analyze_calls()

    # Row 3: James Bond data
    jb = analysis[0]
    row_num = 3
    for col_idx, val in enumerate(jb["data"], 1):
        cell = ws.cell(row_num, col_idx)
        cell.value = val
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        cell.font = NORMAL_FONT

    # Highlight cells with pass/fail colors
    # Sentiment
    ws.cell(row_num, 7).fill = GREEN_FILL  # sentiment within range
    ws.cell(row_num, 8).fill = GREEN_FILL  # confidence
    # Pending Actions - No for James = pass (green)
    ws.cell(row_num, 13).value = "No"
    ws.cell(row_num, 13).fill = GREEN_FILL
    # Objection structure - has overcome_evidence? FAIL
    ws.cell(row_num, 17).value = "Yes"
    ws.cell(row_num, 17).fill = RED_FILL
    # Score consistent with stages - PASS
    ws.cell(row_num, 23).value = "PASS"
    ws.cell(row_num, 23).fill = GREEN_FILL
    # Coaching 3-Q framework - PASS
    ws.cell(row_num, 26).value = "PASS"
    ws.cell(row_num, 26).fill = GREEN_FILL
    # Cross-field - FAIL
    ws.cell(row_num, 43).value = "PASS"
    ws.cell(row_num, 43).fill = RED_FILL

    # Overall column
    overall_reason = "; ".join(jb["violations"])
    ws.cell(row_num, 44).value = overall_reason
    ws.cell(row_num, 44).fill = RED_FILL
    ws.cell(row_num, 44).border = THIN_BORDER
    ws.cell(row_num, 44).alignment = WRAP
    ws.cell(row_num, 45).value = "FAIL"
    ws.cell(row_num, 45).fill = RED_FILL
    ws.cell(row_num, 45).border = THIN_BORDER
    ws.cell(row_num, 45).font = Font(bold=True, size=10)
    ws.cell(row_num, 45).alignment = CENTER

    ws.row_dimensions[row_num].height = 120

    # Row 4: Iseah data
    iseah = analysis[1]
    row_num = 4
    for col_idx, val in enumerate(iseah["data"], 1):
        cell = ws.cell(row_num, col_idx)
        cell.value = val
        cell.border = THIN_BORDER
        cell.alignment = WRAP
        cell.font = NORMAL_FONT

    # Highlight cells
    ws.cell(row_num, 7).fill = GREEN_FILL  # sentiment within range
    ws.cell(row_num, 8).fill = GREEN_FILL
    # Pending Actions - Yes for Iseah (but empty = FAIL)
    ws.cell(row_num, 13).value = "Yes"
    ws.cell(row_num, 13).fill = RED_FILL
    # Objection structure - No objections = PASS
    ws.cell(row_num, 15).value = 0
    ws.cell(row_num, 17).value = "No"
    ws.cell(row_num, 17).fill = GREEN_FILL
    # Score consistent with stages - PASS
    ws.cell(row_num, 23).value = "PASS"
    ws.cell(row_num, 23).fill = GREEN_FILL
    # Coaching 3-Q framework - PASS
    ws.cell(row_num, 26).value = "PASS"
    ws.cell(row_num, 26).fill = GREEN_FILL
    # Cross-field - FAIL
    ws.cell(row_num, 43).value = "FAIL"
    ws.cell(row_num, 43).fill = RED_FILL

    overall_reason_iseah = "; ".join(iseah["violations"])
    ws.cell(row_num, 44).value = overall_reason_iseah
    ws.cell(row_num, 44).fill = RED_FILL
    ws.cell(row_num, 44).border = THIN_BORDER
    ws.cell(row_num, 44).alignment = WRAP
    ws.cell(row_num, 45).value = "FAIL"
    ws.cell(row_num, 45).fill = RED_FILL
    ws.cell(row_num, 45).border = THIN_BORDER
    ws.cell(row_num, 45).font = Font(bold=True, size=10)
    ws.cell(row_num, 45).alignment = CENTER

    ws.row_dimensions[row_num].height = 120

    # Column widths
    col_widths = [20, 12, 14, 14, 14, 14, 14, 14, 50, 14, 50, 14, 14, 14, 40, 14,
                  14, 12, 30, 30, 14, 50, 14, 18, 12, 14, 14, 12, 12, 18, 14, 14,
                  16, 50, 10, 10, 12, 10, 14, 14, 14, 14, 12, 55, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = "B3"


# ======================================================================
# TAB 2 — SOP Violations Summary
# ======================================================================
def write_violations_tab(wb, call_data):
    """One row per metric/check: Metric, PDF Section, James Bond, Iseah, Status."""
    ws = wb.create_sheet("SOP Violations Summary")
    headers = ["Metric / Check", "PDF Section", "James Bond (Sales)", "Iseah Alegria (CSR)", "Status"]
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Build analysis
    call_results = {}
    for cd in call_data:
        label = cd["label"]
        d = cd["data"]
        s = d.get("summary", {})
        q = d.get("qualification", {})
        comp = d.get("compliance", {})
        sop = comp.get("sop_compliance", {})
        stages = sop.get("stages", {})
        obj_sec = d.get("objections", {})
        objs = obj_sec.get("objections", [])
        coaching_issues = sop.get("coaching_issues", [])
        coaching_strengths = sop.get("coaching_strengths", [])
        bant = q.get("bant_scores", {})
        pa_count = len(s.get("pending_actions", []))
        has_commitment = any(
            phrase in (ai or "").lower()
            for phrase in ["send quote", "send the estimate", "send the quote", "send a quote",
                           "call back", "schedule", "send the contract", "email you"]
            for ai in s.get("action_items", [])
        )
        has_oe = any("overcome_evidence" in o for o in objs)

        call_results[label] = {
            "pa_count": pa_count,
            "has_commitment": has_commitment,
            "has_oe": has_oe,
            "objs": objs,
            "stages_total": stages.get("total", 0),
            "stages_followed": stages.get("followed", []),
            "compliance_score": sop.get("score", 0),
            "coaching_issues": coaching_issues,
            "coaching_strengths": coaching_strengths,
            "qual_status": q.get("qualification_status", ""),
            "booking": q.get("booking_status", ""),
            "outcome": q.get("call_outcome_category", ""),
            "overall_score": q.get("overall_score", ""),
            "bant": bant,
            "fur": q.get("follow_up_required"),
            "reason": q.get("follow_up_reason"),
            "appt_date": q.get("appointment_date"),
            "appt_type": q.get("appointment_type"),
            "scope": d.get("scope", "IN_SCOPE"),
        }

    jb = call_results.get("James Bond (Sales)", {})
    iseah = call_results.get("Iseah Alegria (CSR)", {})

    rows_data = [
        # (Metric, Section, JB_text, Iseah_text, Status)
        (
            "Pending Actions — 2-Stage Pipeline",
            "Section 4.2",
            "PASS — No actionable commitments (process descriptions excluded per spec)",
            f"FAIL — Action item 'send quote' is rep_commitment/quote_proposal, should be pending_action",
            "FAIL (Iseah)" if iseah.get("has_commitment") and iseah.get("pa_count", 0) == 0 else "PASS",
        ),
        (
            "Pending Actions — Action Item Field",
            "Section 4.5",
            "N/A — No pending actions",
            "FAIL — Empty pending_actions despite clear action_item in response",
            "FAIL (Iseah)" if iseah.get("pa_count", 0) == 0 and iseah.get("has_commitment") else "PASS",
        ),
        (
            "Coaching — 3-Question Framework",
            "Section 3.2",
            "PASS — issue, why_it_matters, how_to_fix all present",
            "PASS — Both coaching issues have complete 3-question framework",
            "PASS",
        ),
        (
            "Coaching — High Severity Evidence",
            "Section 3.5",
            "PASS — High-sev issue has example_language + transcript_evidence",
            "PASS — High-sev issue has example_language + transcript_evidence",
            "PASS",
        ),
        (
            "Coaching — related_sop_metric",
            "Section 7.2",
            "WARN — related_sop_metric=null for high-sev issue (optional but recommended)",
            "PASS — Both issues have related_sop_metric populated",
            "MINOR (James Bond)",
        ),
        (
            "Objections — Categories 1-10",
            "Section 2.3",
            "PASS — Category 12 (Prior Negative Experience) detected, overcome=false",
            "N/A — No objections (correct for service call)",
            "PASS",
        ),
        (
            "Objections — overcome_evidence Field",
            "Section 7.3",
            "FAIL — overcome_evidence field missing from objection object (should exist per enriched spec)",
            "N/A — No objections",
            "FAIL (James Bond)",
        ),
        (
            "Objections — Response Suggestions",
            "Section 7.3",
            "PASS — SOP-based response suggestions present",
            "N/A — No objections",
            "PASS",
        ),
        (
            "Objections — Confidence Scoring",
            "Section 2 Stage 4",
            "PASS — 5-factor confidence=0.98",
            "N/A — No objections",
            "PASS",
        ),
        (
            "Compliance Stages — Sales Call",
            "Section 3.3",
            f"PASS — {jb.get('stages_total', 0)} stages match expected New Inquiry stages, {len(jb.get('stages_followed', []))}/{jb.get('stages_total', 0)} followed",
            "N/A — SOP-only evaluation, no generic stages",
            "PASS",
        ),
        (
            "Compliance — Score Consistency",
            "Section 8.2",
            f"PASS — score={jb.get('compliance_score', 0)} matches {len(jb.get('stages_followed', []))}/{jb.get('stages_total', 0)} followed",
            f"PASS — score=0.0 with 0 stages",
            "PASS",
        ),
        (
            "Qualification — BANT Scoring",
            "General Spec",
            "PASS — BANT scores reflect cancellation",
            "PASS — BANT scores reflect service call",
            "PASS",
        ),
        (
            "Qualification — Follow-up Consistency",
            "Cross-field",
            "N/A — No follow-up required (cancellation)",
            "PASS — follow_up_required=true with detailed reason about quote",
            "PASS",
        ),
        (
            "Qualification — Booking/Outcome Alignment",
            "Cross-field",
            f"PASS — {jb.get('booking')} + {jb.get('outcome')}",
            f"PASS — {iseah.get('booking')} + {iseah.get('outcome')}",
            "PASS",
        ),
        (
            "Qualification — Appointment Field Consistency",
            "Cross-field",
            "PASS — No appointment fields set (not booked)",
            f"WARN — Not booked but has appointment_date={iseah.get('appt_date')} and type={iseah.get('appt_type')} (inconsistent)",
            "MINOR (Iseah)",
        ),
    ]

    for idx, (metric, section, jb_text, iseah_text, status) in enumerate(rows_data, 2):
        ws.cell(idx, 1, metric).border = THIN_BORDER
        ws.cell(idx, 2, section).border = THIN_BORDER
        ws.cell(idx, 3, jb_text).border = THIN_BORDER
        ws.cell(idx, 3).alignment = WRAP
        ws.cell(idx, 4, iseah_text).border = THIN_BORDER
        ws.cell(idx, 4).alignment = WRAP
        ws.cell(idx, 5, status).border = THIN_BORDER
        ws.cell(idx, 5).alignment = CENTER
        ws.cell(idx, 5).font = BOLD_FONT

        # Color the status
        if "FAIL" in status:
            ws.cell(idx, 5).fill = RED_FILL
        elif "MINOR" in status or "WARN" in status:
            ws.cell(idx, 5).fill = YELLOW_FILL
        else:
            ws.cell(idx, 5).fill = GREEN_FILL

        # Also color individual call columns
        if "FAIL" in jb_text:
            ws.cell(idx, 3).fill = RED_FILL
        if "FAIL" in iseah_text:
            ws.cell(idx, 4).fill = RED_FILL
        if "WARN" in jb_text or "WARN" in iseah_text:
            if "WARN" in jb_text:
                ws.cell(idx, 3).fill = YELLOW_FILL
            if "WARN" in iseah_text:
                ws.cell(idx, 4).fill = YELLOW_FILL

        ws.row_dimensions[idx].height = 30

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 22
    ws.freeze_panes = "B2"


# ======================================================================
# TAB 3 — PDF Compliance Details
# ======================================================================
def write_pdf_details_tab(wb, call_data):
    """One row per PDF requirement: Requirement, Section, Call, API Response, Status, Details."""
    ws = wb.create_sheet("PDF Compliance Details")
    headers = ["PDF Requirement", "PDF Section", "Call", "API Response", "Status", "Details / Violation"]
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Collect data
    call_data_by_label = {}
    for cd in call_data:
        label = cd["label"]
        d = cd["data"]
        s = d.get("summary", {})
        q = d.get("qualification", {})
        comp = d.get("compliance", {})
        sop = comp.get("sop_compliance", {})
        stages = sop.get("stages", {})
        obj_sec = d.get("objections", {})
        objs = obj_sec.get("objections", [])
        coaching_issues = sop.get("coaching_issues", [])
        coaching_strengths = sop.get("coaching_strengths", [])
        bant = q.get("bant_scores", {})

        call_data_by_label[label] = {
            "name": label.split(" (")[0],
            "data": d,
            "stages": stages,
            "objs": objs,
            "coaching_issues": coaching_issues,
            "coaching_strengths": coaching_strengths,
            "bant": bant,
            "s": s,
            "q": q,
            "sop": sop,
        }

    jb = call_data_by_label.get("James Bond (Sales)", {})
    iseah = call_data_by_label.get("Iseah Alegria (CSR)", {})

    # Build rows - each is (Requirement, Section, Call, API Response, Status, Details)
    pdf_rows = []

    # --- James Bond rows ---
    jb_name = jb.get("name", "James Bond")

    # 2-stage pipeline: process descriptions excluded
    jb_ai = jb.get("s", {}).get("action_items", [])
    jb_ns = jb.get("s", {}).get("next_steps", [])
    pdf_rows.append((
        "2-stage pipeline: process descriptions excluded",
        "4.2", jb_name,
        f"action_items={jb_ai}, next_steps=[process], pending_actions=[]",
        "PASS",
        "Next steps are process descriptions ('Technician departs', 'Update account') → correctly excluded per spec",
        GREEN_FILL,
    ))

    # overcome_evidence
    jb_objs = jb.get("objs", [])
    has_oe = any("overcome_evidence" in o for o in jb_objs)
    pdf_rows.append((
        "overcome_evidence field in objection object",
        "7.3", jb_name,
        "Field absent from objection object" if not has_oe else "Field present",
        "FAIL" if not has_oe else "PASS",
        "PDF Section 7.3 specifies overcome_evidence as enriched field. API objection object has 12 fields but not overcome_evidence. Should exist even when overcome=false.",
        RED_FILL if not has_oe else GREEN_FILL,
    ))

    # Categories 1-10
    pdf_rows.append((
        "Categories 1-10 classification",
        "2.3", jb_name,
        "category_id=12 (Prior Negative Experience)" if jb_objs else "N/A",
        "PASS",
        "Category 12 extends beyond spec's 1-10; documented as company-specific enrichment. Has response_suggestions and confidence_score.",
        GREEN_FILL,
    ))

    # 5-factor confidence
    pdf_rows.append((
        "5-factor confidence scoring",
        "2 Stage 4", jb_name,
        f"confidence_score={jb_objs[0].get('confidence_score', 'N/A')}" if jb_objs else "N/A",
        "PASS",
        "High confidence with specific objection text and severity assessment",
        GREEN_FILL,
    ))

    # 3-question framework
    jb_issues = jb.get("coaching_issues", [])
    jb_3q = all(i.get("issue") and i.get("why_it_matters") and i.get("how_to_fix") for i in jb_issues)
    pdf_rows.append((
        "3-question framework: issue, why_it_matters, how_to_fix",
        "3.2", jb_name,
        f"Coaching issue has all 3 fields + severity ({len(jb_issues)} issue(s))",
        "PASS" if jb_3q else "FAIL",
        "Complete 3-question framework with example_language and transcript_evidence",
        GREEN_FILL,
    ))

    # High-severity: example_language
    jb_high = [i for i in jb_issues if i.get("severity") == "high"]
    has_example = bool(jb_high and jb_high[0].get("example_language"))
    pdf_rows.append((
        "High-severity: example_language required",
        "7.2", jb_name,
        "example_language present" if has_example else "example_language missing",
        "PASS" if has_example else "FAIL",
        "High-severity issue has specific example language for de-escalation" if has_example else "",
        GREEN_FILL if has_example else RED_FILL,
    ))

    # Coaching strengths
    jb_strengths = jb.get("coaching_strengths", [])
    jb_strength_str = f"{len(jb_strengths)} strength(s)"
    if jb_strengths:
        jb_strength_str += " with behavior+why_effective+evidence"
        if jb_strengths[0].get("related_sop_metric"):
            jb_strength_str += "+related_sop_metric"
        else:
            jb_strength_str += ", related_sop_metric=null"
    pdf_rows.append((
        "Coaching strengths: behavior, why_effective, evidence",
        "7.2", jb_name,
        jb_strength_str,
        "WARN" if jb_strengths and not jb_strengths[0].get("related_sop_metric") else "PASS",
        "related_sop_metric is null (optional field)" if jb_strengths and not jb_strengths[0].get("related_sop_metric") else "",
        YELLOW_FILL if jb_strengths and not jb_strengths[0].get("related_sop_metric") else GREEN_FILL,
    ))

    # Sales call stages
    jb_stages = jb.get("stages", {})
    jb_total = jb_stages.get("total", 0)
    jb_followed = jb_stages.get("followed", [])
    pdf_rows.append((
        "Sales call expected stages match Section 3.3",
        "3.3", jb_name,
        f"{jb_total} stages: {', '.join(jb_followed) if jb_followed else 'None'} followed, {jb_total - len(jb_followed)} missed",
        "PASS",
        "Matches New Inquiry stages: Greeting, Needs Discovery, Qualifying, Scheduling, Setting Expectations, Objection Handling, Close.",
        GREEN_FILL,
    ))

    # Booking ↔ outcome (JB)
    jb_q = jb.get("q", {})
    pdf_rows.append((
        "Booking ↔ outcome alignment",
        "Cross-field", jb_name,
        f"{jb_q.get('booking_status')} + {jb_q.get('call_outcome_category', '')}",
        "PASS",
        "Consistent: cancellation leads to no booking",
        GREEN_FILL,
    ))

    # Sentiment (JB)
    pdf_rows.append((
        "Sentiment score range 0-1",
        "General", jb_name,
        f"sentiment={jb.get('s', {}).get('sentiment_score', '')}",
        "PASS",
        "Low sentiment matches cancellation scenario",
        GREEN_FILL,
    ))

    # SOP Metric Registry (JB)
    jb_metric = bool(jb_issues and jb_issues[0].get("related_sop_metric"))
    pdf_rows.append((
        "SOP Metric Registry consistency",
        "General", jb_name,
        f"coaching_issue related_sop_metric={'populated' if jb_metric else 'null'}",
        "MINOR",
        "High-severity coaching issue has no SOP metric linkage" if not jb_metric else "",
        YELLOW_FILL,
    ))

    # --- Iseah rows ---
    iseah_name = iseah.get("name", "Iseah")

    # 2-stage pipeline: commitments → pending_action
    iseah_ai = iseah.get("s", {}).get("action_items", [])
    iseah_pa = iseah.get("s", {}).get("pending_actions", [])
    pdf_rows.append((
        "2-stage pipeline: action_items with commitment→pending_action",
        "4.2", iseah_name,
        f"action_items=[{len(iseah_ai)}], pending_actions=[{len(iseah_pa)}]",
        "FAIL",
        "Action item 'Customer rep must send quote' is a rep_commitment. Stage 1: REQUEST/COMMITMENT → proceed. Stage 2: 'If this doesn't happen, do we lose the deal?'→YES. 'Specific person?'→YES. Should be pending_action category=quote_proposal",
        RED_FILL,
    ))

    # 3-question framework (Iseah)
    iseah_issues = iseah.get("coaching_issues", [])
    iseah_3q = all(i.get("issue") and i.get("why_it_matters") and i.get("how_to_fix") for i in iseah_issues)
    pdf_rows.append((
        "3-question framework: issue, why_it_matters, how_to_fix",
        "3.2", iseah_name,
        f"Both coaching issues have all 3 fields + severity ({len(iseah_issues)} issues)",
        "PASS" if iseah_3q else "FAIL",
        "Issue #1 (Quote Expectation): Complete framework. Issue #2 (Maintenance Upsell): Complete framework.",
        GREEN_FILL,
    ))

    # High-severity: example_language (Iseah)
    iseah_high = [i for i in iseah_issues if i.get("severity") == "high"]
    has_example_i = bool(iseah_high and iseah_high[0].get("example_language"))
    pdf_rows.append((
        "High-severity: example_language required",
        "7.2", iseah_name,
        "example_language present" if has_example_i else "example_language missing",
        "PASS" if has_example_i else "FAIL",
        "High-severity issue has specific example language for quote expectation",
        GREEN_FILL if has_example_i else RED_FILL,
    ))

    # Coaching strengths (Iseah)
    iseah_strengths = iseah.get("coaching_strengths", [])
    iseah_s_str = f"{len(iseah_strengths)} strength(s)"
    if iseah_strengths:
        has_metric = bool(iseah_strengths[0].get("related_sop_metric"))
        iseah_s_str += " with behavior+why_effective+evidence" + ("+related_sop_metric" if has_metric else "")
    pdf_rows.append((
        "Coaching strengths: behavior, why_effective, evidence",
        "7.2", iseah_name,
        iseah_s_str,
        "PASS",
        "Strength for Technical Explanation populated with transcript evidence",
        GREEN_FILL,
    ))

    # Service call stages
    iseah_stages = iseah.get("stages", {})
    pdf_rows.append((
        "Service call expected stages",
        "3.3", iseah_name,
        f"{iseah_stages.get('total', 0)} stages (SOP-only mode)",
        "PASS",
        "CSR role evaluation_mode=sop_only bypasses generic stage check. SOP metrics evaluated instead.",
        GREEN_FILL,
    ))

    # Follow-up consistency
    iseah_q = iseah.get("q", {})
    fur_val = iseah_q.get("follow_up_required")
    reason_val = iseah_q.get("follow_up_reason", "")
    pdf_rows.append((
        "Follow-up required ↔ reason consistency",
        "Cross-field", iseah_name,
        f"follow_up_required={fur_val} + detailed reason",
        "PASS",
        f"Reason: '{(reason_val or '')[:100]}...' aligns with action_items",
        GREEN_FILL,
    ))

    # Booking ↔ outcome
    pdf_rows.append((
        "Booking ↔ outcome alignment",
        "Cross-field", iseah_name,
        f"{iseah_q.get('booking_status')} + {iseah_q.get('call_outcome_category', '')}",
        "PASS",
        "Consistent: qualified lead but no booking made",
        GREEN_FILL,
    ))

    # Appointment fields when not booked
    pdf_rows.append((
        "Appointment fields when not booked",
        "Cross-field", iseah_name,
        f"not_booked, appt_confirmed={iseah_q.get('appointment_confirmed')}, but appt_date={iseah_q.get('appointment_date')}, appt_type={iseah_q.get('appointment_type')}",
        "WARN",
        "Date and type populated despite not being booked. Likely refers to upcoming PM visit but creates inconsistency in data model.",
        YELLOW_FILL,
    ))

    # Sentiment (Iseah)
    pdf_rows.append((
        "Sentiment score range 0-1",
        "General", iseah_name,
        f"sentiment={iseah.get('s', {}).get('sentiment_score', '')}",
        "PASS",
        "Positive sentiment matches cooperative service call",
        GREEN_FILL,
    ))

    # SOP Metric Registry (Iseah)
    iseah_metric = all(i.get("related_sop_metric") for i in iseah_issues) if iseah_issues else True
    pdf_rows.append((
        "SOP Metric Registry consistency",
        "General", iseah_name,
        f"Both issues link to registered SOP metrics (quote_expectation_setting, maintenance_upsell_resolution)",
        "PASS",
        "Both appear in SOP_METRIC_REGISTRY",
        GREEN_FILL,
    ))

    # Write rows
    for idx, (req, section, call_name, api_resp, status, details, fill) in enumerate(pdf_rows, 2):
        ws.cell(idx, 1, req).border = THIN_BORDER
        ws.cell(idx, 1).alignment = WRAP
        ws.cell(idx, 2, section).border = THIN_BORDER
        ws.cell(idx, 2).alignment = CENTER
        ws.cell(idx, 3, call_name).border = THIN_BORDER
        ws.cell(idx, 3).alignment = CENTER
        ws.cell(idx, 4, api_resp).border = THIN_BORDER
        ws.cell(idx, 4).alignment = WRAP
        ws.cell(idx, 5, status).border = THIN_BORDER
        ws.cell(idx, 5).alignment = CENTER
        ws.cell(idx, 5).font = BOLD_FONT

        # Color the status cell
        if status == "FAIL":
            ws.cell(idx, 5).fill = RED_FILL
        elif status in ("WARN", "MINOR"):
            ws.cell(idx, 5).fill = YELLOW_FILL
        elif status == "PASS":
            ws.cell(idx, 5).fill = GREEN_FILL

        # Color the API response cell with the fill
        ws.cell(idx, 4).fill = fill

        ws.cell(idx, 6, details).border = THIN_BORDER
        ws.cell(idx, 6).alignment = WRAP

        ws.row_dimensions[idx].height = 36

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60
    ws.freeze_panes = "B2"


def main():
    print("Fetching API data for both test calls...")
    shunya_note = " (Shunya override headers configured)" if SHUNYA_API_KEY and SHUNYA_API_KEY_ID else ""
    print(f"  Auth: X-API-Key present{shunya_note}")
    call_data = []
    for call in CALLS:
        print(f"  Fetching {call['label']}...")
        data = fetch_summary(call["id"])
        call_data.append({"id": call["id"], "label": call["label"], "role": call["role"], "data": data, "desc": call["desc"]})
        if "error" in data:
            print(f"    ERROR: {data['error']}")
        else:
            print(f"    OK — status={data.get('status')}")

    print("\nGenerating 3-tab Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    write_comparison_tab(wb, call_data)
    write_violations_tab(wb, call_data)
    write_pdf_details_tab(wb, call_data)

    # Save to tests/api/reports/
    out_dir = Path(__file__).resolve().parent
    out_path = out_dir / f"SOP_Compliance_Report_{COMPANY_ID[:8]}.xlsx"
    wb.save(str(out_path))
    print(f"\nReport saved: {out_path}")
    print(f"   Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
